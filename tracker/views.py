from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.sites.shortcuts import get_current_site
from django.db import transaction
from django.db.models.functions import TruncYear
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from openpyxl.chart import DoughnutChart, Reference, Series
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from .models import (
    Proses, Tool, SukuCadang, RiwayatStok, ProductionLog, UserProfile, RiwayatPerbaikan
)
from .forms import (
    UserRegisterForm, ToolForm, MaintenanceForm, StokMasukForm,
    StokKeluarForm, ProductionLogForm, PermissionsForm, ProsesForm, StokHistoryFilterForm
)

# Fungsi ini tidak perlu login
def register(request):
    if User.objects.count() == 0:
        User.objects.create_superuser('admin', 'admin@example.com', 'admin')
        admin_user = User.objects.get(username='admin')
        admin_user.userprofile.is_developer = True
        admin_user.userprofile.nama_lengkap = "Developer"
        admin_user.userprofile.save()
        messages.success(request, f'Akun developer "admin" telah dibuat. Silakan login.')
        return redirect('login')

    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False # Buat user tidak aktif
            user.save()

            # Simpan nama lengkap ke profile
            user.userprofile.nama_lengkap = form.cleaned_data.get('nama_lengkap')
            user.userprofile.save()

            # Kirim email aktivasi
            current_site = get_current_site(request)
            mail_subject = 'Aktivasi Akun Anda'
            message = render_to_string('tracker/email_activation_template.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': urlsafe_base64_encode(force_bytes(user.pk)),
                'token': default_token_generator.make_token(user),
            })
            to_email = form.cleaned_data.get('email')
            send_mail(mail_subject, message, 'from@example.com', [to_email])

            return render(request, 'tracker/register_done.html')
    else:
        form = UserRegisterForm()
    return render(request, 'tracker/register.html', {'form': form})

def activate(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, 'Akun berhasil diaktivasi! Anda sekarang bisa login.')
        return redirect('login')
    else:
        messages.error(request, 'Link aktivasi tidak valid!')
        return redirect('login')

@login_required
def daftar_proses(request):
    if request.method == 'POST':
        form = ProsesForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proses baru berhasil ditambahkan!')
            return redirect('daftar_proses')
    form = ProsesForm()
    semua_proses = Proses.objects.all()
    context = {'semua_proses': semua_proses, 'proses_form': form}
    return render(request, 'tracker/proses_list.html', context)

@login_required
def hapus_proses(request, proses_id):
    # Ambil objek proses yang ingin dihapus, jika tidak ada, tampilkan error 404
    proses = get_object_or_404(Proses, pk=proses_id)

    # Pastikan request adalah POST untuk keamanan
    if request.method == 'POST':
        # Simpan nama proses untuk pesan notifikasi sebelum dihapus
        nama_proses = proses.nama
        proses.delete()
        messages.success(request, f'Proses "{nama_proses}" dan semua tool di dalamnya telah berhasil dihapus.')

    # Arahkan kembali pengguna ke halaman daftar proses
    return redirect('daftar_proses')

@login_required
def daftar_tool(request, proses_id):
    proses = get_object_or_404(Proses, pk=proses_id)
    tools = Tool.objects.filter(proses=proses)
    return render(request, 'tracker/tool_list.html', {'proses': proses, 'tools': tools})

@login_required
def tambah_tool(request):
    if request.method == 'POST':
        form = ToolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('daftar_proses')
    else:
        form = ToolForm()
    return render(request, 'tracker/tool_form.html', {'form': form})

@login_required
def edit_tool(request, tool_id):
    tool = get_object_or_404(Tool, pk=tool_id)
    if request.method == 'POST':
        form = ToolForm(request.POST, instance=tool)
        if form.is_valid():
            form.save()
            return redirect('daftar_tool', proses_id=tool.proses.id)
    else:
        form = ToolForm(instance=tool)
    return render(request, 'tracker/tool_form.html', {'form': form, 'tool': tool})

@login_required
def hapus_tool(request, tool_id):
    tool = get_object_or_404(Tool, pk=tool_id)
    proses_id = tool.proses.id
    if request.method == 'POST':
        tool.delete()
    return redirect('daftar_tool', proses_id=proses_id)

@login_required
def send_to_lab(request, tool_id):
    tool = get_object_or_404(Tool, pk=tool_id)
    proses_id = tool.proses.id
    if request.method == 'POST':
        tool.status = 'Perbaikan'
        tool.proses = None
        tool.save()
        RiwayatPerbaikan.objects.create(
            tool=tool,
            status='Sedang Diperbaiki',
            waktu_masuk=timezone.now()
        )
    return redirect('daftar_tool', proses_id=proses_id)

@login_required
def lab_maintenance_list(request):
    # Ambil data untuk filter
    get_year = request.GET.get('year')
    get_month = request.GET.get('month')

    # Daftar tool yang sedang aktif diperbaiki (selalu ditampilkan)
    active_repairs = RiwayatPerbaikan.objects.filter(status='Sedang Diperbaiki').order_by('waktu_masuk')

    gudang_tpm_tools = Tool.objects.filter(status='Gudang TPM').order_by('id_tool')

    # Riwayat perbaikan yang sudah selesai
    history_list = RiwayatPerbaikan.objects.filter(status='Selesai').order_by('-waktu_selesai')

    # Terapkan filter jika ada
    if get_year and get_year.isdigit():
        history_list = history_list.filter(waktu_selesai__year=int(get_year))
    if get_month and get_month.isdigit():
        history_list = history_list.filter(waktu_selesai__month=int(get_month))

    # Siapkan data untuk dropdown filter
    years_with_history = RiwayatPerbaikan.objects.filter(waktu_selesai__isnull=False).annotate(year=TruncYear('waktu_selesai')).values('year').distinct().order_by('-year')

    context = {
        'active_repairs': active_repairs,
        'gudang_tpm_tools': gudang_tpm_tools,
        'history_list': history_list,
        'years_with_history': [d['year'].year for d in years_with_history],
        'months': [
            {'value': 1, 'name': 'Januari'}, {'value': 2, 'name': 'Februari'},
            {'value': 3, 'name': 'Maret'}, {'value': 4, 'name': 'April'},
            {'value': 5, 'name': 'Mei'}, {'value': 6, 'name': 'Juni'},
            {'value': 7, 'name': 'Juli'}, {'value': 8, 'name': 'Agustus'},
            {'value': 9, 'name': 'September'}, {'value': 10, 'name': 'Oktober'},
            {'value': 11, 'name': 'November'}, {'value': 12, 'name': 'Desember'},
        ],
        'selected_year': int(get_year) if get_year else None,
        'selected_month': int(get_month) if get_month else None,
    }
    return render(request, 'tracker/lab_maintenance_list.html', context)

@login_required
def finish_repair(request, tool_id):
    tool = get_object_or_404(Tool, pk=tool_id)
    try:
        riwayat_aktif = RiwayatPerbaikan.objects.get(tool=tool, status='Sedang Diperbaiki')
    except RiwayatPerbaikan.DoesNotExist:
        messages.error(request, 'Catatan perbaikan aktif tidak ditemukan untuk tool ini.')
        return redirect('lab_maintenance_list')

    if request.method == 'POST':
        # Kita gunakan tool sebagai instance agar form tahu field mana yang harus di-render
        form = MaintenanceForm(request.POST, instance=tool)
        if form.is_valid():
            # Update catatan riwayat
            riwayat_aktif.jenis_kerusakan = form.cleaned_data['jenis_kerusakan']
            riwayat_aktif.part_yang_digunakan = form.cleaned_data['part_yang_digunakan']
            riwayat_aktif.dikerjakan_oleh = request.user.username
            riwayat_aktif.waktu_selesai = timezone.now()
            riwayat_aktif.status = 'Selesai'
            riwayat_aktif.save()

            # Ambil tool instance dari form
            tool_instance = form.save(commit=False)
            tool_instance.status = 'Tersedia'

            # --- LOGIKA BARU UNTUK MERESET SHOT ---
            if form.cleaned_data.get('reset_shot_terpakai'):
                tool_instance.shot_terpakai = 0

            tool_instance.save() # Simpan perubahan pada tool
            # Perlu save_m2m jika ada field ManyToMany, tapi kita tidak punya
            return redirect('lab_maintenance_list')
    else:
        form = MaintenanceForm(instance=tool)

    return render(request, 'tracker/finish_repair_form.html', {'form': form, 'tool': tool})

@login_required
def simpan_di_gudang(request, tool_id):
    tool = get_object_or_404(Tool, pk=tool_id)
    
    # Cari catatan perbaikan yang masih aktif untuk tool ini
    try:
        riwayat_aktif = RiwayatPerbaikan.objects.get(tool=tool, status='Sedang Diperbaiki')
    except RiwayatPerbaikan.DoesNotExist:
        messages.error(request, 'Catatan perbaikan aktif tidak ditemukan untuk tool ini.')
        return redirect('lab_maintenance_list')

    if request.method == 'POST':
        # Ubah status tool
        tool.status = 'Gudang TPM'
        tool.save()

        # Selesaikan catatan perbaikan dengan keterangan
        riwayat_aktif.status = 'Selesai' # Status riwayat tetap selesai
        riwayat_aktif.waktu_selesai = timezone.now()
        riwayat_aktif.dikerjakan_oleh = request.user.username
        riwayat_aktif.jenis_kerusakan = riwayat_aktif.jenis_kerusakan or "N/A" # Isi jika kosong
        # Tambahkan catatan khusus di field part yang digunakan
        riwayat_aktif.part_yang_digunakan = f"Tool disimpan di Gudang TPM. Catatan: {riwayat_aktif.part_yang_digunakan or ''}"
        riwayat_aktif.save()

        messages.success(request, f'Tool "{tool.id_tool}" telah dipindahkan ke Gudang TPM.')
        return redirect('lab_maintenance_list')

    # Jika bukan POST, kembalikan ke halaman lab
    return redirect('lab_maintenance_list')

@login_required
@transaction.atomic
def manajemen_stok(request):
    riwayat = RiwayatStok.objects.all().order_by('-waktu')
    filter_form = StokHistoryFilterForm(request.GET)

    # Logika Filter Baru
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
        if start_date:
            riwayat = riwayat.filter(waktu__date__gte=start_date)
        if end_date:
            riwayat = riwayat.filter(waktu__date__lte=end_date)

    if request.method == 'POST':
        if not (request.user.userprofile.is_developer or request.user.userprofile.can_edit_stok):
            messages.error(request, "Anda tidak memiliki izin untuk melakukan aksi ini.")
            return redirect('manajemen_stok')

        if 'submit_masuk' in request.POST:
            form_masuk = StokMasukForm(request.POST)
            if form_masuk.is_valid():
                nama_baru = form_masuk.cleaned_data['nama_barang_baru']
                barang_ada = form_masuk.cleaned_data['barang_yang_ada']
                jumlah = form_masuk.cleaned_data['jumlah']
                suku_cadang = None
                if nama_baru:
                    suku_cadang, created = SukuCadang.objects.get_or_create(nama__iexact=nama_baru.lower(), defaults={'nama': nama_baru})
                elif barang_ada:
                    suku_cadang = barang_ada

                if suku_cadang:
                    suku_cadang.jumlah_stok += jumlah
                    suku_cadang.save()
                    RiwayatStok.objects.create(suku_cadang=suku_cadang, jumlah=jumlah, status='Masuk', nama_pengguna=request.user.username)

                return redirect('manajemen_stok')

        elif 'submit_keluar' in request.POST:
            form_keluar = StokKeluarForm(request.POST)
            if form_keluar.is_valid():
                suku_cadang = form_keluar.cleaned_data['suku_cadang']
                jumlah = form_keluar.cleaned_data['jumlah']
                if suku_cadang.jumlah_stok >= jumlah:
                    suku_cadang.jumlah_stok -= jumlah
                    suku_cadang.save()
                    RiwayatStok.objects.create(suku_cadang=suku_cadang, jumlah=jumlah, status='Keluar', tujuan=form_keluar.cleaned_data['tujuan'], nama_pengguna=request.user.username)
                else:
                    messages.error(request, 'Stok tidak cukup untuk melakukan pengambilan!')
                return redirect('manajemen_stok')

    # Untuk GET request atau jika form tidak valid
    form_masuk = StokMasukForm()
    form_keluar = StokKeluarForm()
    stok_sekarang = SukuCadang.objects.all().order_by('nama')

    context = {
        'form_masuk': form_masuk,
        'form_keluar': form_keluar,
        'stok_sekarang': stok_sekarang,
        'riwayat': riwayat,
        'filter_form': filter_form # Kirim form filter ke template
    }
    return render(request, 'tracker/manajemen_stok.html', context)

@login_required
def halaman_laporan(request):
    proses_list = Proses.objects.all()
    selected_proses_id = request.GET.get('proses_filter')
    tools = Tool.objects.all()
    if selected_proses_id and selected_proses_id.isdigit():
        tools = tools.filter(proses_id=selected_proses_id)
    pie_data_tersedia = tools.filter(status='Tersedia').count()
    pie_data_dipakai = tools.filter(status='Dipakai').count()
    pie_data_perbaikan = tools.filter(status='Perbaikan').count()
    green_count = 0
    yellow_count = 0
    red_count = 0
    tools_for_performance_check = tools.exclude(status='Perbaikan')
    for tool in tools_for_performance_check:
        if tool.performa > 70: green_count += 1
        elif tool.performa >= 30: yellow_count += 1
        else: red_count += 1
    context = {
        'tools': tools, 'all_proses': proses_list, 'selected_proses_id': int(selected_proses_id) if selected_proses_id else None,
        'pie_data_tersedia': pie_data_tersedia, 'pie_data_dipakai': pie_data_dipakai, 'pie_data_perbaikan': pie_data_perbaikan,
        'green_count': green_count, 'yellow_count': yellow_count, 'red_count': red_count, 'perbaikan_count': pie_data_perbaikan,
    }
    return render(request, 'tracker/halaman_laporan.html', context)

def apply_excel_styles(worksheet, min_row=1, max_col=None):
    """
    Kasih style ke worksheet Excel.
    - Header (row pertama) bold + center + fill warna biru.
    - Border tipis ke semua cell.
    - Auto adjust column width.
    """
    if max_col is None:
        max_col = worksheet.max_column

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Styling header
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=min_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Border untuk semua cell
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    # Autofit kolom
    for i, column_cells in enumerate(worksheet.columns):
        if i < max_col:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = max_length + 2

    return worksheet

@login_required
def download_laporan_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="laporan_performa_tools.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Laporan Performa'

    # --- 1. AMBIL DATA & HITUNG RINGKASAN PERFORMA ---
    tools = Tool.objects.all()
    
    # === BAGIAN YANG DIPERBAIKI ADA DI SINI ===
    baik_count = 0
    waspada_count = 0
    kritis_count = 0
    
    # Ambil dulu tool yang akan dicek performanya
    tools_for_performance_check = tools.exclude(status='Perbaikan')
    
    # Hitung manual dengan perulangan for di Python
    for tool in tools_for_performance_check:
        if tool.performa > 70:
            baik_count += 1
        elif tool.performa >= 30:
            waspada_count += 1
        else:
            kritis_count += 1
            
    # Hitungan untuk tool yang sedang perbaikan tetap sama
    perbaikan_count = tools.filter(status='Perbaikan').count()
    # === AKHIR BAGIAN YANG DIPERBAIKI ===

    # --- 2. TULIS DATA RINGKASAN DI KANAN ATAS (sumber data grafik) ---
    summary_data = [
        ("Ringkasan Performa Tool", ""),
        ("Kondisi Baik (> 70%)", baik_count),
        ("Kondisi Waspada (30% - 70%)", waspada_count),
        ("Kondisi Kritis (< 30%)", kritis_count),
        ("Dalam Perbaikan", perbaikan_count),
    ]
    # Tulis data ini mulai dari sel J1
    for i, (label, value) in enumerate(summary_data, start=1):
        worksheet.cell(row=i, column=10, value=label).font = Font(bold=True if i == 1 else False)
        worksheet.cell(row=i, column=11, value=value)

    # --- 3. BUAT & TEMPATKAN GRAFIK DI KIRI ATAS ---
    chart = DoughnutChart()
    chart.title = "Distribusi Kondisi Tool"
    chart.style = 18 

    labels = Reference(worksheet, min_col=10, min_row=2, max_row=5)
    data = Reference(worksheet, min_col=11, min_row=2, max_row=5)   
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    
    chart.height = 10 
    chart.width = 15

    worksheet.add_chart(chart, "A2")

    # --- 4. BUAT TABEL UTAMA DI BAGIAN BAWAH ---
    table_start_row = 18 
    headers = ['ID Tool', 'Tipe Tool', 'Nomor Seri', 'Proses', 'Stasiun', 'Status', 'Max Shot', 'Shot Terpakai', 'Sisa Shot', 'Performa (%)']
    for col, header_title in enumerate(headers, start=1):
        worksheet.cell(row=table_start_row, column=col, value=header_title)

    for row_num, tool in enumerate(tools, start=table_start_row + 1):
        worksheet.cell(row=row_num, column=1, value=tool.id_tool)
        worksheet.cell(row=row_num, column=2, value=tool.tipe_tool)
        worksheet.cell(row=row_num, column=3, value=tool.nomor_seri)
        worksheet.cell(row=row_num, column=4, value=tool.proses.nama if tool.proses else 'Tidak di proses')
        worksheet.cell(row=row_num, column=5, value=tool.stasiun)
        worksheet.cell(row=row_num, column=6, value=tool.status)
        worksheet.cell(row=row_num, column=7, value=tool.max_shot)
        worksheet.cell(row=row_num, column=8, value=tool.shot_terpakai)
        worksheet.cell(row=row_num, column=9, value=tool.sisa_shot)
        worksheet.cell(row=row_num, column=10, value=tool.performa) # Pemanggilan di sini tidak masalah

    # --- 5. TERAPKAN STYLE PADA TABEL UTAMA ---
    apply_excel_styles(worksheet, min_row=table_start_row, max_col=len(headers))
    
    workbook.save(response)
    return response

@login_required
def download_laporan_lab(request):
    get_year = request.GET.get('year')
    get_month = request.GET.get('month')

    riwayat_list = RiwayatPerbaikan.objects.filter(status='Selesai').order_by('-waktu_selesai')

    if get_year and get_year.isdigit():
        riwayat_list = riwayat_list.filter(waktu_selesai__year=int(get_year))
    if get_month and get_month.isdigit():
        riwayat_list = riwayat_list.filter(waktu_selesai__month=int(get_month))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="riwayat_perbaikan_tool.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Riwayat Perbaikan'

    headers = [
        'ID Tool', 'Tipe Tool', 'Nomor Seri', 'Jenis Kerusakan', 'Part yang Digunakan',
        'Dikembalikan ke Proses', 'Tanggal Masuk', 'Waktu Masuk', 'Tanggal Selesai',
        'Waktu Selesai', 'Dikerjakan Oleh'
    ]
    worksheet.append(headers)

    for riwayat in riwayat_list:
        waktu_masuk_local = timezone.localtime(riwayat.waktu_masuk)
        waktu_selesai_local = timezone.localtime(riwayat.waktu_selesai) if riwayat.waktu_selesai else None

        worksheet.append([
            riwayat.tool.id_tool,
            riwayat.tool.tipe_tool,
            riwayat.tool.nomor_seri,
            riwayat.jenis_kerusakan,
            riwayat.part_yang_digunakan,
            riwayat.tool.proses.nama if riwayat.tool.proses else '-',
            waktu_masuk_local.strftime('%Y-%m-%d'),
            waktu_masuk_local.strftime('%H:%M:%S'),
            waktu_selesai_local.strftime('%Y-%m-%d') if waktu_selesai_local else '',
            waktu_selesai_local.strftime('%H:%M:%S') if waktu_selesai_local else '',
            riwayat.dikerjakan_oleh
        ])

    worksheet = apply_excel_styles(worksheet, min_row=1, max_col=worksheet.max_column)
    workbook.save(response)
    return response

@login_required
@transaction.atomic
def halaman_produksi(request):
    if request.method == 'POST':
        form = ProductionLogForm(request.POST)
        if form.is_valid():
            log_entry = form.save()
            jumlah_produksi_hari_ini = log_entry.jumlah_produksi
            durasi_jam = form.cleaned_data.get('durasi_produksi')
            tools_untuk_diupdate = Tool.objects.filter(proses__isnull=False).exclude(status='Perbaikan')

            for tool in tools_untuk_diupdate:
                shot_yang_digunakan = jumlah_produksi_hari_ini * tool.jumlah_shot
                tool.shot_terpakai += shot_yang_digunakan
                tool.jam_pakai_terakumulasi += durasi_jam
                tool.save()

            messages.success(request, f'Data telah ditambahkan ke {tools_untuk_diupdate.count()} tool aktif.')
            return redirect('halaman_produksi')
    else:
        form = ProductionLogForm()

    riwayat_produksi = ProductionLog.objects.all().order_by('-dibuat_pada')[:10]
    context = {'form': form, 'riwayat_produksi': riwayat_produksi}
    return render(request, 'tracker/halaman_produksi.html', context)

@login_required
def hak_akses_list(request):
    if not request.user.userprofile.is_developer:
        return redirect('daftar_proses')
    users = User.objects.exclude(pk=request.user.pk)
    return render(request, 'tracker/hak_akses_list.html', {'users': users})

@login_required
def edit_hak_akses(request, user_id):
    if not request.user.userprofile.is_developer:
        return redirect('daftar_proses')
    user_to_edit = get_object_or_404(User, pk=user_id)
    profile_to_edit = user_to_edit.userprofile
    if request.method == 'POST':
        form = PermissionsForm(request.POST, instance=profile_to_edit)
        if form.is_valid():
            form.save()
            return redirect('hak_akses_list')
    else:
        form = PermissionsForm(instance=profile_to_edit)
    return render(request, 'tracker/edit_hak_akses.html', {'form': form, 'user_to_edit': user_to_edit})

@login_required
def hapus_pengguna(request, user_id):
    # Keamanan: Pastikan hanya developer yang bisa menghapus
    if not request.user.userprofile.is_developer:
        messages.error(request, "Anda tidak memiliki izin untuk melakukan aksi ini.")
        return redirect('daftar_proses')

    # Cari user yang akan dihapus
    user_to_delete = get_object_or_404(User, pk=user_id)

    # Pastikan developer tidak menghapus akunnya sendiri
    if user_to_delete == request.user:
        messages.error(request, "Anda tidak dapat menghapus akun Anda sendiri.")
        return redirect('hak_akses_list')

    if request.method == 'POST':
        username = user_to_delete.username
        user_to_delete.delete()
        messages.success(request, f'Akun "{username}" telah berhasil dihapus.')

    return redirect('hak_akses_list')

@login_required
def get_tool_history(request, tool_id):
    tool = get_object_or_404(Tool, pk=tool_id)
    history = tool.riwayat_perbaikan.all().order_by('-waktu_selesai')
    data = list(history.values('waktu_selesai', 'jenis_kerusakan', 'part_yang_digunakan', 'dikerjakan_oleh'))
    for item in data:
        item['waktu_selesai'] = item['waktu_selesai'].strftime('%d %b %Y, %H:%M')
    return JsonResponse({'history': data})

@login_required
def download_laporan_lab(request):
    get_year = request.GET.get('year')
    get_month = request.GET.get('month')

    riwayat_list = RiwayatPerbaikan.objects.filter(status='Selesai').order_by('-waktu_selesai')

    if get_year and get_year.isdigit():
        riwayat_list = riwayat_list.filter(waktu_selesai__year=int(get_year))
    if get_month and get_month.isdigit():
        riwayat_list = riwayat_list.filter(waktu_selesai__month=int(get_month))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="riwayat_perbaikan_tool.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Riwayat Perbaikan'

    headers = [
        'ID Tool', 'Tipe Tool', 'Nomor Seri', 'Jenis Kerusakan', 'Part yang Digunakan',
        'Dikembalikan ke Proses', 'Tanggal Masuk', 'Waktu Masuk', 'Tanggal Selesai',
        'Waktu Selesai', 'Dikerjakan Oleh'
    ]
    worksheet.append(headers)

    # --- BAGIAN INI DIPERBAIKI TOTAL ---
    # Urutan data sekarang sama persis dengan urutan header di atas
    for riwayat in riwayat_list:
        worksheet.append([
            riwayat.tool.id_tool,
            riwayat.tool.tipe_tool,
            riwayat.tool.nomor_seri,
            riwayat.jenis_kerusakan,
            riwayat.part_yang_digunakan,
            riwayat.tool.proses.nama if riwayat.tool.proses else '-',
            riwayat.waktu_masuk.strftime('%d %b %Y'),
            riwayat.waktu_masuk.strftime('%H:%M'),
            riwayat.waktu_selesai.strftime('%d %b %Y') if riwayat.waktu_selesai else '',
            riwayat.waktu_selesai.strftime('%H:%M') if riwayat.waktu_selesai else '',
            riwayat.dikerjakan_oleh
        ])

    worksheet = apply_excel_styles(worksheet)
    workbook.save(response)
    return response

@login_required
def download_laporan_stok(request):
    riwayat = RiwayatStok.objects.all().order_by('-waktu')
    filter_form = StokHistoryFilterForm(request.GET)
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
        if start_date:
            riwayat = riwayat.filter(waktu__date__gte=start_date)
        if end_date:
            riwayat = riwayat.filter(waktu__date__lte=end_date)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="laporan_stok.xlsx"'
    workbook = openpyxl.Workbook()

    # Sheet 1: Riwayat Stok
    worksheet1 = workbook.active
    worksheet1.title = 'Riwayat Stok'
    headers1 = ['Nama Barang', 'Jumlah', 'Status', 'Tanggal', 'Waktu', 'Tujuan', 'Nama']
    worksheet1.append(headers1)

    for log in riwayat:
        # --- PERBAIKAN DI SINI ---
        # Konversi waktu ke zona waktu lokal (WIB) sebelum dicetak
        waktu_local = timezone.localtime(log.waktu)
        worksheet1.append([
            log.suku_cadang.nama,
            log.jumlah,
            log.status,
            waktu_local.strftime('%Y-%m-%d'),
            waktu_local.strftime('%H:%M:%S'),
            log.tujuan,
            log.nama_pengguna
        ])

    # Sheet 2: Stok Saat Ini
    worksheet2 = workbook.create_sheet(title="Stok Saat Ini")
    headers2 = ['Nama Barang', 'Jumlah Stok Saat Ini']
    worksheet2.append(headers2)
    stok_sekarang = SukuCadang.objects.all().order_by('nama')
    for item in stok_sekarang:
        worksheet2.append([item.nama, item.jumlah_stok])

    worksheet1 = apply_excel_styles(worksheet1)
    worksheet2 = apply_excel_styles(worksheet2)

    workbook.save(response)
    return response
